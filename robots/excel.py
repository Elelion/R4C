import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from .models import Robot  # Импортируйте модель Robot из вашего приложения

def create_robot_summary_excel():
    # Создаем новую рабочую книгу и активный лист
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Robot Summary"

    # Заголовок таблицы
    headers = ["Модель", "Версия", "Количество за неделю"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Получите данные о роботах из базы данных и заполните таблицу
    robots = Robot.objects.all()  # Подставьте ваш запрос к модели Robot
    row_num = 2  # Начинаем с второй строки
    for robot in robots:
        sheet.cell(row=row_num, column=1, value=robot.model)
        sheet.cell(row=row_num, column=2, value=robot.version)
        sheet.cell(row=row_num, column=3, value=robot.quantity)  # Здесь укажите количество роботов за неделю
        row_num += 1

    # Создаем таблицу для данных
    table = Table(displayName="RobotTable", ref=sheet.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True
    )
    table.tableStyleInfo = style
    sheet.add_table(table)

    # Сохраняем файл
    filename = "robot_summary.xlsx"
    workbook.save(filename)
    return filename
